#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
利率曲线数据更新脚本 (本地版 — 引用自 yyseles/bond-yield-curve)
四条曲线：国债即期 / 国开债即期 / 国债到期 / 国开债到期
"""
import json, os, sys, tempfile, time
from datetime import datetime, date, timedelta, timezone
import requests
from openpyxl import load_workbook

DATA_FILE = "data.json"
CDB_DATA_FILE = "data_cdb.json"
GOV_YTM_FILE = "data_gov_ytm.json"
CDB_YTM_FILE = "data_cdb_ytm.json"
SUMMARY_FILE = "summary.json"
VERSION_FILE = "version.json"

CHINABOND_DOWNLOAD_URL = "https://yield.chinabond.com.cn/cbweb-mn/yc/bxjDownload"
SEARCHYC_URL = "https://yield.chinabond.com.cn/cbweb-mn/yc/searchYc"
GOV_CURVE_ID = "2c9081e50a2f9606010a3068cae70001"
CDB_CURVE_ID = "8a8b2ca037a7ca910137bfaa94fa5057"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Referer": "https://yield.chinabond.com.cn/cbweb-mn/yc/bxjInit?locale=zh_CN",
}
SEARCHYC_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Referer": "https://yield.chinabond.com.cn/cbweb-mn/yield_main?locale=zh_CN",
    "Content-Type": "application/x-www-form-urlencoded",
}

ALL_TERMS = [f"{i}Y" for i in range(1, 51)]
SUMMARY_TERMS = ["1Y", "5Y", "10Y", "20Y", "30Y"]
BJ_TZ = timezone(timedelta(hours=8))
MAX_RETRIES = 3
RETRY_DELAY = 5

def now_beijing():
    return datetime.now(BJ_TZ).date()

# ============ 国债即期 (XLSX下载) ============
def fetch_spot_rates_chinabond(query_date):
    params = {"gzr": query_date, "csz": "1", "locale": "zh_CN"}
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.post(CHINABOND_DOWNLOAD_URL, params=params, headers=HEADERS, timeout=30)
            r.raise_for_status()
            if len(r.content) < 200:
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                    continue
                return {}
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(r.content)
                tmp_path = tmp.name
            try:
                wb = load_workbook(tmp_path)
                ws = wb.active
                data = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[1] is not None and row[2] is not None:
                        data[float(row[1])] = float(row[2])
                wb.close()
            finally:
                os.unlink(tmp_path)
            result = {}
            for y in range(1, 51):
                val = data.get(float(y))
                if val is not None:
                    result[f"{y}Y"] = round(val, 8)
            return result if result else {}
        except Exception as e:
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
            else:
                print(f"  {query_date}: 请求失败 - {e}")
    return {}

# ============ searchYc 通用抓取 ============
def fetch_searchyc_rates(curve_id, qxll, query_date, label=""):
    params = {
        "xyzSelect": "txy", "workTimes": query_date, "dxbj": "0",
        "qxll": qxll, "yqqxN": "N", "yqqxK": "K",
        "ycDefIds": curve_id, "wrjxCBFlag": "0", "locale": "zh_CN",
    }
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.post(SEARCHYC_URL, data=params, headers=SEARCHYC_HEADERS, timeout=30)
            r.raise_for_status()
            data = r.json()
            if not data or not isinstance(data, list):
                if attempt < MAX_RETRIES:
                    time.sleep(RETRY_DELAY)
                    continue
                return {}
            series = data[0].get("seriesData", [])
            result = {}
            for tenor, val in series:
                if abs(tenor - round(tenor)) < 1e-6 and 1 <= tenor <= 50:
                    result[f"{int(tenor)}Y"] = round(val, 8)
            return result if result else {}
        except Exception as e:
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY)
            else:
                print(f"  [{label}] {query_date}: 请求失败 - {e}")
    return {}

# ============ 文件读写 ============
def load_existing(filepath):
    if not os.path.exists(filepath):
        return {"dates": [], "terms": ALL_TERMS, "rows": []}
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    if len(data.get("terms", [])) < 50:
        data["terms"] = ALL_TERMS
    return data

def save_json(filepath, data):
    tmp = filepath + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, filepath)

# ============ 更新函数 ============
def update_curve(name, fetch_fn, data_file, today_str):
    print(f"\n{'─'*40}")
    print(f" [{name}] 开始更新")
    existing = load_existing(data_file)
    print(f"  现有: {len(existing['dates'])} 条")
    
    fetch_start = "2020-01-01"
    if existing["dates"]:
        last = existing["dates"][-1]
        fetch_start = (datetime.strptime(last, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    print(f"  范围: {fetch_start} → {today_str}")
    
    all_new = {}
    current = datetime.strptime(fetch_start, "%Y-%m-%d")
    end = datetime.strptime(today_str, "%Y-%m-%d")
    fetched, skipped = 0, 0
    while current <= end:
        ds = current.strftime("%Y-%m-%d")
        if current.weekday() < 5:
            rates = fetch_fn(ds)
            if rates:
                all_new[ds] = rates
                fetched += 1
                if fetched % 50 == 0:
                    print(f"  已获取 {fetched} 个交易日...")
            else:
                skipped += 1
        current += timedelta(days=1)
    print(f"  获取: {fetched} 个交易日, 跳过/无数据: {skipped} 天")
    
    if not all_new:
        print(f" ⚠ [{name}] 没有新数据")
        return False
    
    date_to_row = {}
    for i, d in enumerate(existing["dates"]):
        date_to_row[d] = existing["rows"][i]
    for d in sorted(all_new.keys()):
        date_to_row[d] = [all_new[d].get(t) for t in ALL_TERMS]
    
    sorted_dates = sorted(date_to_row.keys())
    sorted_rows = [date_to_row[d] for d in sorted_dates]
    
    output = {"dates": sorted_dates, "terms": ALL_TERMS, "rows": sorted_rows}
    save_json(data_file, output)
    print(f" ✅ [{name}] 总计 {len(sorted_dates)} 条")
    return True

# ============ summary.json ============
def generate_summary():
    print(f"\n{'─'*40}")
    print(" [summary] 生成摘要")
    curves = [
        ("gov_spot", DATA_FILE, "国债即期"),
        ("gov_ytm", GOV_YTM_FILE, "国债到期"),
        ("cdb_spot", CDB_DATA_FILE, "国开债即期"),
        ("cdb_ytm", CDB_YTM_FILE, "国开债到期"),
    ]
    summary = {"curves": {}}
    all_dates = set()
    for key, fp, label in curves:
        if not os.path.exists(fp):
            continue
        data = load_existing(fp)
        if not data["dates"]:
            continue
        ld = data["dates"][-1]
        lr = data["rows"][-1]
        terms_data = {}
        for term in SUMMARY_TERMS:
            if term in data["terms"]:
                idx = data["terms"].index(term)
                terms_data[term] = {"value": lr[idx] if idx < len(lr) else None, "change": None}
        summary["curves"][key] = {"name": label, "date": ld, "terms": terms_data}
        all_dates.add(ld)
    summary["date"] = max(all_dates) if all_dates else ""
    save_json(SUMMARY_FILE, summary)
    print(f" ✅ [summary] 日期: {summary['date']}")

# ============ 主函数 ============
def main():
    print("=" * 55)
    print(" 利率曲线 · 本地自动更新")
    print(f" 北京时间: {datetime.now(BJ_TZ).strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 55)
    today_str = now_beijing().strftime("%Y-%m-%d")
    
    ok1 = update_curve("国债即期", fetch_spot_rates_chinabond, DATA_FILE, today_str)
    ok2 = update_curve("国开即期", lambda d: fetch_searchyc_rates(CDB_CURVE_ID, "1", d, "国开即期"), CDB_DATA_FILE, today_str)
    ok3 = update_curve("国债到期", lambda d: fetch_searchyc_rates(GOV_CURVE_ID, "0", d, "国债到期"), GOV_YTM_FILE, today_str)
    ok4 = update_curve("国开到期", lambda d: fetch_searchyc_rates(CDB_CURVE_ID, "0", d, "国开到期"), CDB_YTM_FILE, today_str)
    
    generate_summary()
    
    # 如果有新数据，推送到 GitHub Pages
    if any([ok1, ok2, ok3, ok4]):
        # 生成新版本号（只有数据真变了才更新，避免无意义缓存失效）
        version_str = datetime.now(BJ_TZ).strftime('%Y%m%d%H%M%S')
        save_json(VERSION_FILE, {"version": version_str, "date": today_str})
        print("\n────────────────────────────────────────")
        print(" [Git] 推送更新到 GitHub Pages")
        import subprocess
        script_dir = os.path.dirname(os.path.abspath(__file__))
        try:
            subprocess.run(["git", "add", DATA_FILE, CDB_DATA_FILE, GOV_YTM_FILE, CDB_YTM_FILE, SUMMARY_FILE, VERSION_FILE],
                          cwd=script_dir, check=True, capture_output=True)
            subprocess.run(["git", "commit", "-m", f"数据更新至{today_str}"],
                          cwd=script_dir, check=True, capture_output=True)
            subprocess.run(["git", "push", "origin", "main"],
                          cwd=script_dir, check=True, capture_output=True, timeout=60)
            print(" ✅ [Git] 推送成功")
        except subprocess.CalledProcessError as e:
            # git commit 在无变更时返回非零，正常
            if "nothing to commit" in (e.stderr.decode() if e.stderr else ""):
                print(" ⚠ [Git] 无新变更需要推送")
            else:
                print(f" ❌ [Git] 推送失败: {e}")
    
    print(f"\n{'='*55}")
    results = [f"国债即期{'✅' if ok1 else '⚠'}", f"国开即期{'✅' if ok2 else '⚠'}",
               f"国债到期{'✅' if ok3 else '⚠'}", f"国开到期{'✅' if ok4 else '⚠'}"]
    print(f" 汇总: {' | '.join(results)}")
    print("=" * 55)

if __name__ == "__main__":
    main()
